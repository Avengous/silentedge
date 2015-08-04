from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime
from LoLPlayer import Player

"""
Sheet: Full
- All Matches
- Important Match Data
- Day by day improvement
- all time champion stats
"""


class Report(Player):
	def __init__(self, playerId):
		self.wb = Workbook()
		self.ws = self.wb.active
		self.ws.title = 'Full'
		self.player = Player(playerId)
		self.headers =  '''matchCreation
									winner
									championId
									champLevel
									kills
									assists
									deaths
									kda
									minionsKilled
									goldEarned
									lane
									matchDuration
									totalDamageDealtToChampions
									totalDamageTaken
									wardsPlaced
									wardsKilled
									visionWardsBoughtInGame
									csDiffPerMin@10
									csDiffPerMin@20
									csDiffPerMinAverage
									goldPerMin@10
									goldPerMin@20
									goldPerMinAverage
									xpDiffPerMin@10
									xpDiffPerMin@20
									xpDiffPerMinAverage
									creepsPerMin@10
									creepsPerMin@20
									creepsPerMinAverage
									xpPerMin@10
									xpPerMin@20
									xpPerMinAverage
									damageTakenDiffPerMin@10
									damageTakenDiffPerMin@20
									damageTakenDiffPerMinAverage
									damageTakenPerMin@10
									damageTakenPerMin@20
									damageTakenPerMinAverage
									item0
									item1
									item2
									item3
									item4
									item5
									item6
									matchHistoryUri'''
	
	def _createCell(self, column, row, value, BOLD=False, WRAP=False):
		self.ws.cell(column=column, row=row).font = Font(bold=BOLD)
		self.ws.cell(column=column, row=row, value=value)
		self.ws.cell(column=column, row=row).alignment = Alignment(wrap_text=WRAP)

	def _createHeaders(self):
		for i, header in enumerate(self.headers.splitlines()):
			self._createCell(i+1, 1, header.strip(), BOLD=True, WRAP=True)
	
	def _createFullRow(self, row, match):
		stats = match.stats
		for col, header in enumerate(self.headers.splitlines()):
			if header.strip() == 'matchCreation':
				date = datetime.fromtimestamp(stats[header.strip()]/1000)
				self._createCell(col+1, row, date)
			elif header.strip() == 'winner':
				if stats[header.strip()] == 1:
					self._createCell(col+1, row, 'WIN')
				else:
					self._createCell(col+1, row, 'LOSS')
			elif header.strip() == 'championId':
				champion = match.champion()
				self._createCell(col+1, row, champion)
			elif header.strip() == 'matchDuration':
				self._createCell(col+1, row, round(stats[header.strip()]/60.0, 2))
			elif header.strip() == 'kda':
				self._createCell(col+1, row, match.kda())
			else:
				try:
					self._createCell(col+1, row, stats[header.strip()])
				except: 
					if 'Average' in header.strip():
						deltas = []
						deltas.append(stats[header.strip().replace('Average', '_zeroToTen')])
						deltas.append(stats[header.strip().replace('Average', '_tenToTwenty')])
						deltas.append(stats[header.strip().replace('Average', '_twentyToThirty')])
						deltas.append(stats[header.strip().replace('Average', '_thirtyToEnd')])
						average = 0
						i = 0
						for v in deltas:
							if v:
								i += 1
								average += v
						try:
							average = average/i
						except:
							average = 0
						self._createCell(col+1, row, round(average, 2))
					else:
						delta = stats[header.strip().replace('@10', '_zeroToTen') \
																	 .replace('@20', '_tenToTwenty')]
						if type(delta) is float:
							delta = round(delta, 2)
						self._createCell(col+1, row, delta)
				
	def create(self):
		self._createHeaders()
		matches = self.player.getMatches()
		
		for i, match in enumerate(matches):
			self._createFullRow(i+2, match)

		column = {}
		for c in self.ws.rows:
			for i in c:
				column[i.column] = 0
			break

		for a, c in enumerate(self.ws.rows):
			if a > 0:
				for b, i in enumerate(c):
					if b>0:
						v = len(str(i.value))*2
						if v > column[i.column]:
							column[i.column] = v
							self.ws.column_dimensions[i.column].width = v
					else:
						self.ws.column_dimensions[i.column].width = 18
			else:
				self.ws.row_dimensions[c[1].row].height = 30
				
		dest_filename =  'LoLReport_{}.xlsx'.format(self.player.playerId)
		self.wb.save(filename=dest_filename)
		
"""
match.stats['matchCreation']
match.stats['magicDamageDealtToChampions']
match.stats['physicalDamageDealtToChampions']
match.stats['trueDamageDealtToChampions']
match.stats['totalDamageDealtToChampions']
match.stats['magicDamageTaken']
match.stats['trueDamageTaken']
match.stats['physicalDamageTaken']
match.stats['totalDamageTaken']
match.stats['towerKills']
match.stats['totalTimeCrowdControlDealt']
match.stats['neutralMinionsKilledEnemyJungle']
match.stats['neutralMinionsKilledTeamJungle']
match.stats['neutralMinionsKilled']
match.stats['firstTowerAssist']
match.stats['firstTowerKill']
match.stats['firstBloodKill']
match.stats['firstBloodAssist']
match.stats['firstInhibitorKill']
match.stats['firstInhibitorAssist']
match.stats['item0']
match.stats['item1']
match.stats['item2']
match.stats['item3']
match.stats['item4']
match.stats['item5']
match.stats['item6']
match.stats['winner']
match.stats['wardsPlaced']
match.stats['wardsKilled']
match.stats['visionWardsBoughtInGame']
match.stats['minionsKilled']
match.stats['champLevel']
match.stats['kills']
match.stats['assists']
match.stats['deaths']
match.stats['goldEarned']
match.stats['inhibitorKills']
match.stats['goldSpent']
match.stats['championId']
match.stats['lane']
match.stats['teamId']
match.stats['spell1Id']
match.stats['spell2Id']
match.stats['role']
match.stats['matchDuration']
match.stats['matchHistoryUri']
match.stats['csDiffPerMin_zeroToTen']
match.stats['csDiffPerMin_tenToTwenty']
match.stats['csDiffPerMin_twentyToThirty']
match.stats['csDiffPerMin_thirtyToEnd']
match.stats['goldPerMin_zeroToTen']
match.stats['goldPerMin_tenToTwenty']
match.stats['goldPerMin_twentyToThirty']
match.stats['goldPerMin_thirtyToEnd']
match.stats['xpDiffPerMin_zeroToTen']
match.stats['xpDiffPerMin_tenToTwenty']
match.stats['xpDiffPerMin_twentyToThirty']
match.stats['xpDiffPerMin_thirtyToEnd']
match.stats['creepsPerMin_zeroToTen']
match.stats['creepsPerMin_tenToTwenty']
match.stats['creepsPerMin_twentyToThirty']
match.stats['creepsPerMin_thirtyToEnd']
match.stats['xpPerMin_zeroToTen']
match.stats['xpPerMin_tenToTwenty']
match.stats['xpPerMin_twentyToThirty']
match.stats['xpPerMin_thirtyToEnd']
match.stats['damageTakenDiffPerMin_zeroToTen']
match.stats['damageTakenDiffPerMin_tenToTwenty']
match.stats['damageTakenDiffPerMin_twentyToThirty']
match.stats['damageTakenDiffPerMin_thirtyToEnd']
match.stats['damageTakenPerMin_zeroToTen']
match.stats['damageTakenPerMin_tenToTwenty']
match.stats['damageTakenPerMin_twentyToThirty']
match.stats['damageTakenPerMin_thirtyToEnd']
"""